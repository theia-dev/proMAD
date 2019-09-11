import io
import json
import shutil
import unittest
import warnings
from pathlib import Path

from helper import hash_file, hash_string_mem
from openpyxl import load_workbook

from proMAD import ArrayAnalyse


class TestWithARY022B(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.cases = Path(__file__).absolute().resolve().parent / 'cases'
        cls.aa = ArrayAnalyse.load(cls.cases / 'save' / 'dump.tar')
        cls.out_folder = cls.cases / 'testing_reports'
        cls.out_folder.mkdir(exist_ok=True, parents=True)

    @classmethod
    def tearDownClass(cls):
        del cls.aa
        shutil.rmtree(cls.out_folder)

    def test_wrong_type(self):
        with warnings.catch_warnings(record=True) as w:
            warnings.simplefilter("always")

            self.aa.report()

            self.assertEqual(len(w), 1)
            self.assertEqual(w[-1].category, RuntimeWarning)
            self.assertIn("is not defined as report type.", str(w[-1].message))

    def test_no_file(self):
        with warnings.catch_warnings(record=True) as w:
            warnings.simplefilter("always")

            self.aa.report(report_type='csv')

            self.assertEqual(len(w), 1)
            self.assertEqual(w[-1].category, RuntimeWarning)
            self.assertIn("No file was given.", str(w[-1].message))

    def test_csv(self):
        test_file_path = self.out_folder / 'rep.csv'
        self.aa.report(test_file_path)
        self.assertEqual(hash_file(test_file_path, skip=0),
                         'da519d1d722cecd6f30a409389e2d141733c58507e81549c67a619cb287c6953')
        test_file_path.unlink()
        self.aa.report(test_file_path, norm='raw')
        self.assertEqual(hash_file(test_file_path, skip=0),
                         '3c093be302781733aadbe063f4e09032a5bf780613a87e780af0b8487efdf5be')

        save_mem = io.StringIO()
        self.aa.report(save_mem, report_type='csv')
        self.assertEqual(hash_string_mem(save_mem),
                         'da519d1d722cecd6f30a409389e2d141733c58507e81549c67a619cb287c6953')

    def test_json(self):
        test_file_path = self.out_folder / 'rep.json'

        compare_11 = {'position': [0, 11], 'info': ['Angiopoietin-2', '285', 'Ang-2, ANGPT2'],
                      'value': 0.9658152124642082, 'intercept': -0.00014130440936884273,
                      'r_squared': 0.9974079209437973}
        compare_123 = {'position': [5, 11], 'info': ['IL-31', '386653', None], 'value': 0.9974935542164649,
                       'intercept': -0.0008484813336183776, 'r_squared': 0.9992305292185427}

        self.aa.report(test_file_path, norm='raw')
        data = json.loads(test_file_path.read_text())

        self.assertAlmostEqual(data['result'][11]['values'][0], 0.023836107062511658, places=7)
        self.assertAlmostEqual(data['result'][11]['values'][2], 0.02761170374937431, places=7)
        self.assertAlmostEqual(data['result'][11]['values'][4], 0.032503605013469966, places=7)

        self.assertAlmostEqual(data['result'][123]['values'][0], 0.023919083139474023, places=7)
        self.assertAlmostEqual(data['result'][123]['values'][2], 0.027834284233911134, places=7)
        self.assertAlmostEqual(data['result'][123]['values'][4], 0.03295996329826092, places=7)

        save_mem = io.StringIO()
        self.aa.report(save_mem, report_type='json')
        save_mem.seek(0)
        data = json.loads(save_mem.read())

        self.assertAlmostEqual(data['result'][11]['value'], compare_11['value'], places=7)
        self.assertAlmostEqual(data['result'][123]['value'], compare_123['value'], places=7)
        self.assertAlmostEqual(data['result'][11]['intercept'], compare_11['intercept'], places=7)
        self.assertAlmostEqual(data['result'][123]['intercept'], compare_123['intercept'], places=7)
        self.assertAlmostEqual(data['result'][11]['r_squared'], compare_11['r_squared'], places=7)
        self.assertAlmostEqual(data['result'][123]['r_squared'], compare_123['r_squared'], places=7)

    def test_excel(self):
        test_file_path = self.out_folder / 'rep.xls'
        self.aa.report(test_file_path)
        test_file_path = test_file_path.with_suffix('.xlsx')
        wb_file = load_workbook(filename=test_file_path, read_only=True)

        save_mem = io.BytesIO()
        self.aa.report(save_mem, report_type='excel')
        wb_mem = load_workbook(filename=save_mem, read_only=True)

        for wb in [wb_mem, wb_file]:
            ws = wb['Overview']
            self.assertAlmostEqual(ws['B4'].value, 3.58838189472233, places=7)
            self.assertAlmostEqual(ws['Y4'].value, 2.8227990771837, places=7)
            self.assertAlmostEqual(ws['B13'].value, 3.72477602867501, places=7)
            self.assertEqual(ws['B15'].value, 'Fig. 1: Overview and alignment check')
            self.assertEqual('A1:Y15', ws.calculate_dimension())

            ws = wb['Results']
            self.assertEqual('A1:C112', ws.calculate_dimension())
            self.assertEqual(ws['A4'].value, 'Serpin E1')
            self.assertEqual(ws['B4'].value, '[8, 0]')
            self.assertAlmostEqual(ws['C4'].value, 3.83146659901802, places=7)

            ws = wb['Result details']
            self.assertEqual('A1:E221', ws.calculate_dimension())
            self.assertEqual(ws['E3'].value, 'R_Squared')
            self.assertEqual(ws['B4'].value, '[8, 0]')
            self.assertEqual(ws['B7'].value, '[8, 1]')
            self.assertAlmostEqual(ws['C4'].value, 4.12145253276224, places=7)
            self.assertAlmostEqual(ws['C7'].value, 3.54148066527379, places=7)

            ws = wb['Info']
            self.assertEqual('A1:B15', ws.calculate_dimension())
            info_data = []
            for row in ws.rows:
                for cell in row:
                    info_data.append(cell.value)
            self.assertIn('ARY022B', info_data)
            self.assertIn('hist_raw', info_data)
            self.assertIn('Norm description', info_data)


if __name__ == '__main__':
    unittest.main()